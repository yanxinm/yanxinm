#!/usr/bin/env python3
"""
Batch convert Windows Office documents (.docx, .xlsx, .pdf) to Markdown (.md)
for import into GBrain knowledge base.

Usage:
    python3 convert_docs.py --src "/mnt/e/百度云同步盘/工作台账" \\
                            --dst ~/brain_src/工作台账 \\
                            --years 2021-2026

Output: Mirror directory structure under --dst with .md files only.
Idempotent: skips files where .md is newer than source.
"""
import os
import subprocess
import sys
import argparse
from pathlib import Path


def parse_args():
    parser = argparse.ArgumentParser(description="Convert Office docs to Markdown")
    parser.add_argument("--src", required=True, help="Source directory (Windows path)")
    parser.add_argument("--dst", required=True, help="Destination directory for .md files")
    parser.add_argument("--years", default=None, help="Comma-separated or hyphen-range years, e.g. '2021-2026' or '2025,2026'")
    return parser.parse_args()


def expand_years(spec: str) -> list[str]:
    """Parse '2021-2026' or '2025,2026' into a list of year strings."""
    if not spec:
        return []
    result = []
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-", 1)
            result.extend(str(y) for y in range(int(start), int(end) + 1))
        else:
            result.append(part)
    return result


def convert_docx(src_path: str, dst_md: str) -> None:
    """Convert .docx to .md via pandoc."""
    result = subprocess.run(
        ["pandoc", src_path, "-t", "markdown", "--wrap=preserve", "-o", dst_md],
        capture_output=True, text=True, timeout=60
    )
    if result.returncode != 0:
        raise RuntimeError(f"pandoc error: {result.stderr[:200]}")


def convert_pdf(src_path: str, dst_md: str) -> None:
    """Extract PDF text via pymupdf."""
    try:
        import pymupdf as fitz
    except ImportError:
        import fitz
    doc = fitz.open(src_path)
    text_lines = []
    for page in doc:
        text_lines.append(page.get_text())
    doc.close()
    with open(dst_md, "w", encoding="utf-8") as f:
        f.write("\n\n".join(text_lines))


def convert_xlsx(src_path: str, dst_md: str) -> None:
    """Convert .xlsx to markdown tables via openpyxl.
    Falls back to xlrd for legacy .xls format (Excel 97-2003)."""
    import openpyxl
    ext = os.path.splitext(src_path)[1].lower()

    if ext == ".xls":
        # Legacy .xls — try xlrd, fall back gracefully
        try:
            import xlrd
            wb = xlrd.open_workbook(src_path)
            parts = []
            for sheet_name in wb.sheet_names():
                ws = wb.sheet_by_name(sheet_name)
                parts.append(f"## {sheet_name}\n")
                for i in range(min(ws.nrows, 81)):
                    vals = [str(ws.cell_value(i, c)) if ws.cell_value(i, c) != "" else "" for c in range(ws.ncols)]
                    parts.append("| " + " | ".join(vals) + " |\n")
                    if i == 0:
                        parts.append("| " + " | ".join(["---"] * ws.ncols) + " |\n")
                if ws.nrows > 81:
                    parts.append("_（更多行已省略）_\n")
                parts.append("\n")
            with open(dst_md, "w", encoding="utf-8") as f:
                f.writelines(parts)
            return
        except ImportError:
            raise RuntimeError(".xls format requires xlrd: pip install xlrd")
        except Exception as e:
            raise RuntimeError(f"xlrd error: {e}")

    # .xlsx — use openpyxl
    wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"## {sheet_name}\n")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 80:
                parts.append("_（更多行已省略）_\n")
                break
            vals = [str(v) if v is not None else "" for v in row]
            parts.append("| " + " | ".join(vals) + " |\n")
            if i == 0:
                parts.append("| " + " | ".join(["---"] * len(vals)) + " |\n")
        parts.append("\n")
    wb.close()
    with open(dst_md, "w", encoding="utf-8") as f:
        f.writelines(parts)


def main():
    args = parse_args()
    src_base = os.path.abspath(args.src)
    dst_base = os.path.abspath(args.dst)
    years = expand_years(args.years) if args.years else []

    if not os.path.isdir(src_base):
        print(f"[错误] 源目录不存在: {src_base}", file=sys.stderr)
        sys.exit(1)

    os.makedirs(dst_base, exist_ok=True)
    stats = {"docx": 0, "pdf": 0, "xlsx": 0, "skipped": 0, "errors": 0}
    total = 0

    # Determine which top-level dirs to process
    if years:
        dirs_to_process = []
        for y in years:
            p = os.path.join(src_base, y)
            if os.path.isdir(p):
                dirs_to_process.append(p)
    else:
        dirs_to_process = [d for d in Path(src_base).iterdir() if d.is_dir() and not d.name.startswith(".")]

    for src_dir in dirs_to_process:
        year_or_name = os.path.basename(src_dir)
        print(f"\n=== 处理 {year_or_name} ===")
        for root, dirs, files in os.walk(src_dir):
            # Skip hidden dirs
            dirs[:] = [d for d in dirs if not d.startswith(".")]
            for fname in files:
                if fname.startswith(".") or fname.startswith("~$"):
                    stats["skipped"] += 1
                    continue

                src_file = os.path.join(root, fname)
                rel = os.path.relpath(src_file, src_base)
                dst_file = os.path.join(dst_base, rel)
                dst_dir = os.path.dirname(dst_file)
                dst_md = dst_file + ".md"

                # Skip if .md exists and is newer than source
                if os.path.isfile(dst_md):
                    if os.path.getmtime(dst_md) >= os.path.getmtime(src_file):
                        stats["skipped"] += 1
                        continue

                os.makedirs(dst_dir, exist_ok=True)

                try:
                    lower = fname.lower()
                    if lower.endswith(".docx"):
                        convert_docx(src_file, dst_md)
                        stats["docx"] += 1
                    elif lower.endswith(".pdf"):
                        convert_pdf(src_file, dst_md)
                        stats["pdf"] += 1
                    elif lower.endswith((".xlsx", ".xls")):
                        convert_xlsx(src_file, dst_md)
                        stats["xlsx"] += 1
                    else:
                        stats["skipped"] += 1
                        continue

                    total += 1
                    if total % 20 == 0:
                        print(f"  已处理 {total} 个文件...")
                except Exception as e:
                    stats["errors"] += 1
                    print(f"  [错误] {rel}: {e}", file=sys.stderr)

    print(f"\n===== 转换完成 =====")
    print(f"  docx → md: {stats['docx']}")
    print(f"  pdf → md:  {stats['pdf']}")
    print(f"  xlsx → md: {stats['xlsx']}")
    print(f"  跳过:      {stats['skipped']}")
    print(f"  错误:      {stats['errors']}")
    print(f"  总计转换:  {total} 个文件")


if __name__ == "__main__":
    main()
